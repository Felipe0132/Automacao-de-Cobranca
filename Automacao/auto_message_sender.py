from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from urllib.parse import quote
from datetime import datetime
import tempfile
import time
import customtkinter as ctk

data_atual = datetime.now()

#FUNÇOES PARA INTERFACE

ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue") 

# Criar janela principal

janela = ctk.CTk()
janela.title("Automatizador de Mensagens")
janela.geometry("400x500")

janela_inicial = ctk.CTkFrame(janela, fg_color="white") # Cor de fundo
janela_inicial.pack(fill="both", expand=True) # Cria um widget, fill é o preenchimento, expand seria a responsividade do tamanho

janela_dados = ctk.CTkFrame(janela, fg_color="white") # Cor da segunda janela


def voltar():
    janela_dados.pack_forget() # Fecha a janela
    janela_inicial.pack(fill="both", expand=True)

#FUNÇOES PARA ABRIR WHATSAPP

options = Options()
options.binary_location = "/usr/bin/google-chrome"  # Caminho do Chrome no Linux Mint
options.add_argument("--start-maximized") # Navegador em tela cheia
temp_dir = tempfile.mkdtemp()
options.add_argument(f"--user-data-dir={temp_dir}") # Abre novo navegadpr
options.add_argument("--profile-directory=Default") # Aqui cria um usuario temporario para não conflitar com um Chrome já aberto 

# Inicializa o ChromeDriver automaticamente
service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=service, options=options)

driver.get("https://web.whatsapp.com")

#FUNÇOES PARA LER PLANILHA


# Lendo a planilha
planilha_cobranca = load_workbook('Automacao/cobranca.xlsx')

# Lendo as paginas da planilha
paginas = planilha_cobranca.sheetnames  
print(paginas)

# Criando a lista de paginas
paginas_planilha = []

# Criando dicionario para identificadores especificos
lista_devedores = list()
lista_credores = list()
lista_cliente = list()

# Adicionando paginas a lista de paginas para manipulaçao
for i in range(len(paginas)):
    paginas_planilha.append(planilha_cobranca[paginas[i]])


# FUNÇOES PROPRIAS

def mostrando_clientes():
    janela_popup = ctk.CTkToplevel(janela) # Cria uma janela secundaria da principal
    janela_popup.title("Lista de Clientes") 
    janela_popup.geometry("400x400")
    janela_popup.configure(fg_color="white")

    janela_lista = ctk.CTkScrollableFrame(janela_popup, fg_color="white") # fundo
    janela_lista.pack(fill="both", expand=True, padx=10, pady=10)

    titulo_devedores = ctk.CTkLabel(
        janela_lista,
        text="Devedores:",
        text_color="black",
        font=("Arial", 24, "bold"),
        anchor="w", # Posição que o texto ficará, west
        fg_color="transparent" # Fundo 
    )

    titulo_devedores.pack(pady=(5, 5), anchor="w")

    for infos in lista_devedores:
        texto = f"{infos['Nome']}, situação R${infos['Valor']}"
        ctk.CTkLabel(janela_lista, text=texto, text_color="black", anchor="w").pack(pady=2, fill="x") # fill -> Faz espandir pelo x

    separador = ctk.CTkLabel(
        janela_lista,
        text="",
        fg_color="#E0E0E0",
        height=1
    )

    separador.pack(fill="x", pady=10)

    titulo_credores = ctk.CTkLabel(
        janela_lista,
        text="Credores:",
        text_color="black",
        font=("Arial", 24, "bold"),
        anchor="w",
        fg_color="transparent"
    )

    titulo_credores.pack(pady=(5, 5), anchor="w")

    for infos in lista_credores:
        texto = f"{infos['Nome']}, situação R${infos['Valor']}"
        ctk.CTkLabel(janela_lista, text=texto, text_color="black", anchor="w").pack(pady=2, fill="x")

    ctk.CTkButton(janela_popup, text="Fechar", command=janela_popup.destroy, fg_color="#B0B0B0", text_color="black").pack(pady=10) # Botão de voltar e acabar com janela secundaria

def buscar_cliente(): # Necessita nome identico aos que estao presentes
    janela_popup = ctk.CTkToplevel(janela)
    janela_popup.title("Buscando Cliente")
    janela_popup.geometry("400x400")
    janela_popup.configure(fg_color="white")

    titulo = ctk.CTkLabel(janela_popup, text="Buscando Cliente", text_color="black", font=("Arial", 22))
    titulo.pack(pady=10)

    entrada_nome = ctk.CTkEntry(janela_popup, placeholder_text="Digite o nome do cliente", width=350)
    entrada_nome.pack(pady=10)

    janela_lista = ctk.CTkScrollableFrame(janela_popup, fg_color="white")
    janela_lista.pack(fill="both", expand=True, padx=10, pady=10)

    def executar_busca():
        nome = entrada_nome.get()

        for widget in janela_lista.winfo_children()[1:]: # Aqui ele seleciona o segundo widget que há na janela, no caso os dados
            widget.destroy()  # Limpa resultados anteriores

        if nome == "":
            ctk.CTkLabel(janela_lista, text="Campo vazio!", text_color="red").pack(pady=2)
            return


        for infos in lista_cliente:            
            if nome != infos['Nome']:
                continue
            else:
                if infos['Valor'] > 0:
                    msg = f'devendo R${infos['Valor']}'
                elif infos['Valor'] < 0:
                    msg = f'credito R${abs(infos['Valor'])}'
                else:
                    msg = f'R${infos['Valor']}'

                texto = f"Nome: {infos['Nome']}\nValor em aberto: {msg}\nContato: {infos['Contato']}"
                ctk.CTkLabel(janela_lista, text=texto, text_color="black", justify="left").pack(pady=5, anchor="w")
                return
                
        ctk.CTkLabel(janela_lista, text="Cliente não encontrado!", text_color="red").pack(pady=2)


    # Cria o campo dos botões abaixo, para não conflitar com o espaço
    janela_botoes = ctk.CTkFrame(janela_popup, fg_color="white")
    janela_botoes.pack(pady=10)

    btn_buscar = ctk.CTkButton(
        janela_botoes,
        text="Buscar",
        command=executar_busca,
        fg_color="#808080",
        text_color="white"
    )
    btn_buscar.pack(side="left", padx=10)

    btn_voltar = ctk.CTkButton(
        janela_botoes,
        text="Voltar",
        command=janela_popup.destroy,
        fg_color="#4682B4",
        text_color="white"
    )
    btn_voltar.pack(side="left", padx=10)


def mostrar_devedores():
    janela_popup = ctk.CTkToplevel(janela)
    janela_popup.title("Lista de Devedores")
    janela_popup.geometry("400x400")
    janela_popup.configure(fg_color="white")

    titulo = ctk.CTkLabel(janela_dados, text="Lista de Devedores", text_color="black", font=("Arial", 22))
    titulo.pack(pady=10)

    janela_lista = ctk.CTkScrollableFrame(janela_popup, fg_color="white")
    janela_lista.pack(fill="both", expand=True, padx=10, pady=10)

    titulo_devedores = ctk.CTkLabel(
        janela_lista,
        text="Devedores:",
        text_color="black",
        font=("Arial", 24, "bold"),
        anchor="w",
        fg_color="transparent"
    )

    titulo_devedores.pack(pady=(5, 5), anchor="w")

    for infos in lista_devedores:
        texto = f"{infos['Nome']}, situação R${infos['Valor']}"
        ctk.CTkLabel(janela_lista, text=texto, text_color="black", anchor="w").pack(pady=2, fill="x")

    ctk.CTkButton(janela_popup, text="Fechar", command=janela_popup.destroy, fg_color="#B0B0B0", text_color="black").pack(pady=10)

def mostrar_ndevedores():
    janela_popup = ctk.CTkToplevel(janela)
    janela_popup.title("Lista de Credores")
    janela_popup.geometry("400x400")
    janela_popup.configure(fg_color="white")

    titulo = ctk.CTkLabel(janela_dados, text="Lista de Credores", text_color="black", font=("Arial", 22))
    titulo.pack(pady=10)

    janela_lista = ctk.CTkScrollableFrame(janela_popup, fg_color="white")
    janela_lista.pack(fill="both", expand=True, padx=10, pady=10)

    titulo_credores = ctk.CTkLabel(
        janela_lista,
        text="Credores:",
        text_color="black",
        font=("Arial", 24, "bold"),
        anchor="w",
        fg_color="transparent"
    )

    titulo_credores.pack(pady=(5, 5), anchor="w")

    for infos in lista_credores:
        texto = f"{infos['Nome']}, situação R${abs(infos['Valor'])}"
        ctk.CTkLabel(janela_lista, text=texto, text_color="black", anchor="w").pack(pady=2, fill="x")

    ctk.CTkButton(janela_popup, text="Fechar", command=janela_popup.destroy, fg_color="#B0B0B0", text_color="black").pack(pady=10)

def mensagem_todos():
    for cliente in lista_cliente:

        msg = f"Bom dia {cliente['Nome']}, tudo bem? Passando para atualizar sobre sua situação conosco,"

        if cliente['Valor'] >= 0:
            msg += f" você está com R${cliente['Valor']} em aberto!"
        else:
            msg += f" você está com R${abs(cliente['Valor'])} de crédito!"

        numero = cliente['Contato']
        if not numero.startswith('+'):
            numero = '+55' + numero

        print(f"Enviando mensagem para {cliente['Nome']} ({numero})...")

        msg = quote(msg)

        try:
            driver.get(f"https://web.whatsapp.com/send?phone={numero}&text={msg}")
            chat_box = WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.XPATH, "//footer//div[@role='textbox']"))
            )

            # Dá foco no campo e envia ENTER
            chat_box.click()
            time.sleep(1)
            chat_box.send_keys(Keys.ENTER) 
            # Ele espera o campo de texto estar ok para assim efetivar o ENTER

            print(f"Mensagem enviada para {cliente['Contato']} com sucesso!\n")
            time.sleep(5)

            janela_popup = ctk.CTkToplevel(janela)
            janela_popup.title("Aviso")
            janela_popup.geometry("400x400")
            janela_popup.configure(fg_color="white")

            msg = f'Mensagem enviada para {cliente['Contato']} com sucesso!\n'

            texto = ctk.CTkLabel(janela_dados, text="{msg}", text_color="black", font=("Arial", 18))
            texto.pack(pady=10)

        except Exception as e:
            print(f"Erro ao enviar mensagem para {cliente['Nome']}: {e}")

    msg = 'Ação Finalizada!'

    texto = ctk.CTkLabel(janela_dados, text="{msg}", text_color="black", font=("Arial", 18))
    texto.pack(pady=10)

    ctk.CTkButton(janela_popup, text="Fechar", command=janela_popup.destroy, fg_color="#B0B0B0", text_color="black").pack(pady=10)   

def mensagem_devedores():
    for cliente in lista_devedores:

        msg = f"Bom dia {cliente['Nome']}, tudo bem? Passando para atualizar sobre sua situação conosco, você está com R$ {cliente['Valor']} em aberto!"

        numero = cliente['Contato']
        if not numero.startswith('+'):
            numero = '+55' + numero

        print(f"Enviando mensagem para {cliente['Nome']} ({numero})...")

        try:
            driver.get(f"https://web.whatsapp.com/send?phone={numero}&text={msg}")
            chat_box = WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.XPATH, "//footer//div[@role='textbox']"))
            )

            # Dá foco no campo e envia ENTER
            chat_box.click()
            time.sleep(1)
            chat_box.send_keys(Keys.ENTER)

            print(f"Mensagem enviada para {cliente['Contato']} com sucesso!\n")
            time.sleep(5)

            janela_popup = ctk.CTkToplevel(janela)
            janela_popup.title("Aviso")
            janela_popup.geometry("400x400")
            janela_popup.configure(fg_color="white")

            msg = f'Mensagem enviada para {cliente['Contato']} com sucesso!\n'

            texto = ctk.CTkLabel(janela_dados, text="{msg}", text_color="black", font=("Arial", 18))
            texto.pack(pady=10)
        except Exception as e:
            print(f"Erro ao enviar mensagem para {cliente['Nome']}: {e}")

    msg = 'Ação Finalizada!'

    texto = ctk.CTkLabel(janela_dados, text="{msg}", text_color="black", font=("Arial", 18))
    texto.pack(pady=10)

    ctk.CTkButton(janela_popup, text="Fechar", command=janela_popup.destroy, fg_color="#B0B0B0", text_color="black").pack(pady=10)   


def mensagem_informacoes():

    janela_popup = ctk.CTkToplevel(janela)
    janela_popup.title("Enviando informações")
    janela_popup.geometry("400x400")
    janela_popup.configure(fg_color="white")

    titulo = ctk.CTkLabel(janela_popup, text="Enviando Mensagem", text_color="black", font=("Arial", 22))
    titulo.pack(pady=10)
    titulo = ctk.CTkLabel(janela_popup, text="Digite seu número:", text_color="black", font=("Arial", 17))
    titulo.pack(pady=10)

    entrada_nome = ctk.CTkEntry(janela_popup, placeholder_text="+55 necessario")
    entrada_nome.pack(pady=10)

    janela_lista = ctk.CTkScrollableFrame(janela_popup, fg_color="white")
    janela_lista.pack(fill="both", expand=True, padx=10, pady=10)

    def enviar():

        msg = f'Lista de devedores:\n'
        for cliente in lista_devedores:
            msg += f'{cliente['Nome']}. devendo: R${cliente['Valor']}\n'

        msg += f'\nLista de Credores:\n'
        for cliente in lista_credores:
            msg += f'{cliente['Nome']}. situação: R${abs(cliente['Valor'])}\n'

        try:
            numero_user = entrada_nome.get()

            driver.get(f"https://web.whatsapp.com/send?phone={numero_user}&text={msg}")
            chat_box = WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.XPATH, "//footer//div[@role='textbox']"))
            )

            # Dá foco no campo e envia ENTER
            chat_box.click()
            time.sleep(1)
            chat_box.send_keys(Keys.ENTER)

            print(f"Mensagem enviada para seu contato com sucesso!\n")
            time.sleep(5)

            msg = f'Mensagem enviada para {cliente['Contato']} com sucesso!\n'

            texto = ctk.CTkLabel(janela_dados, text="{msg}", text_color="black", font=("Arial", 18))
            texto.pack(pady=10)
        except Exception as e:
            print(f"Erro ao enviar mensagem para seu contato: {e}")

        msg = 'Ação Finalizada!'

    texto = ctk.CTkLabel(janela_dados, text="{msg}", text_color="black", font=("Arial", 18))
    texto.pack(pady=10)

        # Cria o campo dos botões abaixo, para não conflitar com o espaço
    janela_botoes = ctk.CTkFrame(janela_popup, fg_color="white")
    janela_botoes.pack(pady=10)

    btn_buscar = ctk.CTkButton(
        janela_botoes,
        text="Enviar",
        command=enviar,
        fg_color="#808080",
        text_color="white"
    )
    btn_buscar.pack(side="left", padx=10)

    btn_voltar = ctk.CTkButton(
        janela_botoes,
        text="Voltar",
        command=janela_popup.destroy,
        fg_color="#4682B4",
        text_color="white"
    )
    btn_voltar.pack(side="left", padx=10)

paginas_planilha[1]['B3'].value = -100


# Colocando clientes nas listas adequadas
for i in range(len(paginas)):
    #print(f"\n----- {paginas[i]} -----")
    for indice, linha in enumerate(paginas_planilha[i].iter_rows(values_only=True)):

        if indice == 0:
            nome = linha[1]
        
        if indice == 1:
            contato = str(linha[1])
        
        if indice == 2:
            valor = linha[1]

            dados = {'Nome' : nome, 'Valor' : valor, 'Contato' : contato}

            lista_cliente.append(dados)

            if valor > 0:
                    #print(f"{nome} devendo R${valor}!")

                    lista_devedores.append(dados)

            else:
                    #print(f"{nome} nao esta devendo!")

                    lista_credores.append(dados)
            

"""
print("\nLista de devedores:")
for nome, valor in lista_devedores.items():
    print(f"{nome} R${valor}", end=", ")

print()

print("\nLista de credores:")      
for nome, valor in lista_credores.items():
    print(f"{nome} R${valor}", end=", ") 
"""


titulo = ctk.CTkLabel(
    janela_inicial,
    text="Automatizador de Mensagens",
    font=("Arial Rounded MT Bold", 20),
    text_color="black",
    fg_color="transparent"
)
titulo.pack(pady=(20, 20))

botao1 = ctk.CTkButton(janela, text="Mostrar todos os clientes", command=mostrando_clientes, width=270, height=40)
botao1.pack(pady=5)

botao2 = ctk.CTkButton(janela, text="Mostrar clientes devedores", command=mostrar_devedores, width=270, height=40)
botao2.pack(pady=5)

botao3 = ctk.CTkButton(janela, text="Mostrar clientes nao devedores", command=mostrar_ndevedores, width=270, height=40)
botao3.pack(pady=5)

botao4 = ctk.CTkButton(janela, text="Buscar cliente", command=buscar_cliente, width=270, height=40) 
botao4.pack(pady=5)

botao5 = ctk.CTkButton(janela, text="Mandar mensagem para todos clientes", command=mensagem_todos, width=270, height=40)
botao5.pack(pady=5)

botao6 = ctk.CTkButton(janela, text="Mandar mensagem para devedores", command=mensagem_devedores, width=270, height=40)
botao6.pack(pady=5)

botao7 = ctk.CTkButton(janela, text="Mandar informacoes para seu Whatsapp", command=mensagem_informacoes, width=270, height=40)
botao7.pack(pady=5)

botao8 = ctk.CTkButton(janela, text="Sair", command=janela.quit, width=270, height=40)
botao8.pack(pady=5)

    
janela.mainloop()

driver.quit()